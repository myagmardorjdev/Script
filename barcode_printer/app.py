current_dir = "C:/Users/myagmardorj/Git/lesson3/barcode_printer/"
from flask import *
from barcode_print import *
import io
import psycopg2
import openpyxl
import xlwings as xw
from datetime import datetime
import pypyodbc as odbc

odoodatabases = {'34': {'user': 'readonly_c34','password': 'readonly_c34_password','server': '10.34.1.220','port': 5432,'database':'CARREFOURS34_LIVE'},
                '13': {'user': 'postgres','password': 'postgres','server': '10.13.1.220','port': 5432,'database':'CARREFOURS13_LIVE'},
                '88': {'user': 'readonly_c88','password': 'readonly_c88_password','server': '10.88.1.220','port': 5432,'database':'CARREFOURS88_LIVE'},
                '21': {'user': 'readonly_c21','password': 'readonly_c21_password','server': '10.21.1.220','port': 5432,'database':'CARREFOURS21_LIVE'},
                '01': {'user': 'readonly_c01','password': 'readonly_c01_password','server': '10.1.1.220',  'port': 5432,'database':'STORE01_LIVE'},
                '12': {'user': 'readonly_c42','password': 'readonly_c42_password','server': '10.12.1.220',  'port': 5432,'database':'STORE42_LIVE'},
                '06': {'user': 'readonly_c06','password': 'readonly_c06_password','server': '10.6.1.220',  'port': 5432,'database':'STORE06_LIVE'},
                '17': {'user': 'readonly_c17','password': 'readonly_c17_password','server': '10.17.1.220',  'port': 5432,'database':'STORE17_LIVE'},
                '38': {'user': 'readonly_c38','password': 'readonly_c38_password','server': '10.38.1.220',  'port': 5432,'database':'STORE38_LIVE'},
                '25': {'user': 'readonly_c25','password': 'readonly_c25_password','server': '10.25.1.220',  'port': 5432,'database':'STORE25_LIVE'},
                '26': {'user': 'readonly_c26','password': 'readonly_c26_password','server': '10.26.1.220',  'port': 5432,'database':'STORE26_LIVE'},
                '16': {'user': 'readonly_c16','password': 'readonly_c16_password','server': '10.16.1.220',  'port': 5432,'database':'STORE16_LIVE'},
                '22': {'user': 'readonly_c22','password': 'readonly_c22_password','server': '10.22.1.220',  'port': 5432,'database':'STORE22_LIVE'},
                '08': {'user': 'readonly_c08','password': 'readonly_c08_password','server': '10.8.1.220',  'port': 5432,'database':'STORE08_LIVE'},
                '39': {'user': 'readonly_c39','password': 'readonly_c39_password','server': '10.39.1.220',  'port': 5432,'database':'STORE39_LIVE'},
                '32': {'user': 'itid','password': 'it#2016','server': '10.32.1.220',  'port': 5432,'database':'STORE32_LIVE'},
                '61': {'user': 'itid','password': 'it#2016','server': '10.61.1.220',  'port': 5432,'database':'STORE61_LIVE'},
                '62': {'user': 'itid','password': 'it#2016','server': '10.62.1.220',  'port': 5432,'database':'STORE62_LIVE'},
                '33': {'user': 'itid','password': 'it#2016','server': '10.33.1.220',  'port': 5432,'database':'STORE33_LIVE'},
                '20': {'user': 'itid','password': 'it#2016','server': '10.20.1.220',  'port': 5432,'database':'DC'},
                '00': {'user': 'itid','password': 'it#2016','server': '10.0.99.220',  'port': 5432,'database':'CENTRAL_PREGOLIVE'}
}
def postg_return_value(conn,query):
    global stringa;
    cursor = conn.cursor()
    cursor.execute(query)

    data = cursor.fetchall()
    #print(len(data))
    #for row in data:
        #print(row)
        #stringa = row;
    conn.close()
    return data

query_get_barcodes = "SELECT * FROM product_product where barcode = '"
query_get_template = "select name,list_price from product_template where id = '"
query_get_extra_price = "SELECT date_start,date_end,fixed_price,min_quantity,pricelist_id FROM product_pricelist_item where active='true' and (date_end >= CURRENT_DATE or date_end is null) and product_tmpl_id = '"
app = Flask(__name__, static_url_path='/static')
class excel_file_to_list_xlwings(): #! excel file iig list bolgon huwirganga (convert)
    def __init__(self,path,filename,barcodeint):
        self.path = path
        self.filename = filename
        self.barcode = barcodeint
        ws = xw.Book(self.path +self.filename).sheets['Sheet1']
        v1 = ws.range("A2:T1030").value
        self.newlist=[]
        for i in v1:
            if i is not None:
                if(int(i[1])==self.barcode):
                    print("barcode olow")
                    self.newlist.append(i)
                    break
        # v2 = ws.range("F5").value
        self.newlist = i
    def returnc(self):
        return self.newlist

class excel_file_to_list(): #! excel file iig list bolgon huwirganga (convert)
    def __init__(self,path,filename,barcode):
        self.path = path
        self.filename = filename
        self.exceldata = (openpyxl.load_workbook(self.path+self.filename)).active
        self.returnlist = []
        for row in range(1, self.exceldata.max_row):
            self.newrow = []
            for col in self.exceldata.iter_cols(1, self.exceldata.max_column):
                self.newrow.append(col[row].value)
            if self.newrow[1] == int(barcode):
                break;
            
    def returnc(self):
        return self.newrow
@app.route('/')
def index():

    return render_template('about.html')

@app.route('/button_clicked', methods=['POST'])
def button_clicked():
    global product_template_table
    global product_pricelist_table
    global product_product_table
    global begindate
    global mainprice
    user_input = request.form.get('user_input')
    ortslist = []
    baraaner = request.form.get('second_input')
    expire = request.form.get('fourthinput')
    begindate = "x"
    min_quantity = 0
    logo = request.form.get('dropdown')
    button_type = request.form.get('button_type')
    ishavedate = request.form.get('dropdown2')
    print('date ',ishavedate)
    papersize = request.form.get('dropdown3')
    if papersize =='6141':
        ortslist = excel_file_to_list_xlwings(current_dir,"data.xlsx",int(user_input)).returnc()
    if logo == "sansar":
        isSansar = 1
    else:
        isSansar = 0

    if ishavedate == "yes":
        ishavedate = 1
    else:
        ishavedate = 0
    if papersize =="4030":
        width_mm = 60   
        height_mm = 45
    elif papersize == "6040":
        width_mm = 90   
        height_mm = 60
    elif papersize =="297210":
        width_mm = 297
        height_mm = 210
    elif papersize =="14876":
        width_mm = 148 
        height_mm = 76
    elif papersize == "6141":
        width_mm = 2400   
        height_mm = 1600
    if button_type == "button1":
        print(product_pricelist_table)
        if papersize =="14876":
            if len(product_pricelist_table) > 0:
                for i in product_pricelist_table:

                    if int(i[3]) != 0 : # 1 ширхэг барааг гэдэг үгийг шалгаж байна
                        print(int(i[3]))
                        now = datetime.now() # хэрэгжих хугацааг одоо цагаас өнгөрсөн байгаа үгүйг шалгаж байна
                        if i[0] is not None and i[4] not in (25,26):
                            if now >= i[0]:
                                print("Бөөний үнэ хэрэгжсэн байна E1")
                                min_quantity = int(i[3])
                                mainprice = int(i[2])   
                    if int(i[3]) != 0 and i[0] is None and i[4] not in (25,26): 
                        print ("Бөөний үнэ хэрэгжсэн байна E2")
                        min_quantity = int(i[3])
                        mainprice = int(i[2])    
            else:
                mainprice = int(product_template_table[0][1])
                min_quantity = 1
        else:
            mainprice = int(product_template_table[0][1])

        innercode = product_product_table[0][14]
        barcode_usr = request.form.get('user_input')
        
        output_file = "output_with_text.png"
        
        text_content = product_template_table[0][0]
        barcode_text = "Код: "+str(barcode_usr)
        barcode = str(barcode_usr)
        defaultprice = int(product_template_table[0][1])
        #print(product_pricelist_table) # 8 цагийн зөрүүтэй байгааг анхаарна уу

        if len(product_pricelist_table) > 0:
            for i in product_pricelist_table:

                if int(i[3]) == 0: # 1 ширхэг барааг гэдэг үгийг шалгаж байна
                    print(int(i[3]))
                    now = datetime.now() # хэрэгжих хугацааг одоо цагаас өнгөрсөн байгаа үгүйг шалгаж байна
                    if i[0] is not None:
                        if now >= i[0]:
                            print("limited үнэ хэрэгжсэн байна E3")
                            defaultprice = int(i[2])
                            begindate = str(i[0])[:10] +";"+ str(str(i[1])[:10])[-5:]
                if int(i[3]) == 0 and i[0] is None:
                    print ("infinity үнэ хэрэгжсэн байна E4")
                    defaultprice = int(i[2])    
                    begindate = str(i[0])[:10] +";"+ str(str(i[1])[:10])[-5:]
        price = "Үнэ: "+str(defaultprice)
    
        image = generate_white_png_with_text(width_mm, height_mm, output_file, text_content,barcode,barcode_text,price,ishavedate,isSansar,expire,innercode,mainprice,begindate,min_quantity,ortslist)
        img_byte_array = io.BytesIO()
        image.save(img_byte_array, format='PNG')
        img_byte_array.seek(0)

        # Return the PNG image as a response with appropriate headers
        
        return send_file(img_byte_array, as_attachment=True, download_name='sample_image.png', mimetype='image/png')
    elif button_type == "button2":
        
        barcode_usr = request.form.get('user_input')
        print("barcode: ",barcode_usr)
        print("date: ",datetime.now())
        itemname_usr = request.form.get('second_input')
        thirtinput_usr = request.form.get('thirtinput')
        salbar_usr=request.form.get('dropdown1')
        papersize = request.form.get('dropdown3')
        print("Цаасны төрөл ", papersize)
        ishavedate = request.form.get('dropdown2')
        type=request.form.get('dropdown')
        expire = request.form.get('fourthinput')
        try:
            conp = psycopg2.connect(database=odoodatabases[salbar_usr[6:]]['database'], user=odoodatabases[salbar_usr[6:]]['user'], password=odoodatabases[salbar_usr[6:]]['password'], host=odoodatabases[salbar_usr[6:]]['server'], port= odoodatabases[salbar_usr[6:]]['port'])
            product_product_table = postg_return_value(conp,query_get_barcodes+barcode_usr+"'")
            
            conp = psycopg2.connect(database=odoodatabases[salbar_usr[6:]]['database'], user=odoodatabases[salbar_usr[6:]]['user'], password=odoodatabases[salbar_usr[6:]]['password'], host=odoodatabases[salbar_usr[6:]]['server'], port= odoodatabases[salbar_usr[6:]]['port'])
            product_template_table = postg_return_value(conp,query_get_template+str(product_product_table[0][4])+"'")
            conp = psycopg2.connect(database=odoodatabases[salbar_usr[6:]]['database'], user=odoodatabases[salbar_usr[6:]]['user'], password=odoodatabases[salbar_usr[6:]]['password'], host=odoodatabases[salbar_usr[6:]]['server'], port= odoodatabases[salbar_usr[6:]]['port'])
            product_pricelist_table = postg_return_value(conp,query_get_extra_price+str(product_product_table[0][4])+"'")
            defaultprice = int(product_template_table[0][1])
            print(product_pricelist_table) # 8 цагийн зөрүүтэй байгааг анхаарна уу
            if len(product_pricelist_table) > 0:
                for i in product_pricelist_table:
                    
                    if int(i[3]) == 0: # 1 ширхэг барааг гэдэг үгийг шалгаж байна
                        now = datetime.now() # хэрэгжих хугацааг одоо цагаас өнгөрсөн байгаа үгүйг шалгаж байна
                        if i[0] is not None:
                            if now >= i[0]:
                                print("үнэ хэрэгжсэн байна")
                                defaultprice = int(i[2])
 
                    if int(i[3]) == 0 and i[0] is None:
                        print ("үнэ хэрэгжих ёстой")
                        defaultprice = int(i[2])    
            defaultvalue = product_template_table[0][0]   # барааны нэр 
            
        except:
            defaultvalue = ""   # барааны нэр 
            defaultprice = -1
            

        
        return render_template('about.html',item_name=defaultvalue,price=defaultprice,barcode=barcode_usr,selected_option=type,selected_salbar=salbar_usr,selected_option2=ishavedate,selected_option3=papersize,expireday=expire)


if __name__ == '__main__':
    app.run(debug=True)